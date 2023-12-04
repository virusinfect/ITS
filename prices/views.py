# views.py
import re
from decimal import Decimal, ROUND_DOWN

import openpyxl
import pandas as pd
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, get_object_or_404
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt

from .forms import PriceListForm
from .models import Supplier, Equipment, Brand, Exchange, LaptopPriceList, ColoursoftPriceList, FellowesPricelist, \
    PriceRule, Type
from .pricing_logic import calculate_discounted_price, calculate_discounted_price2


@login_required
def upload_price_list(request):
    suppliers = Supplier.objects.all()
    brands = Brand.objects.all().order_by("name")
    equipment = Equipment.objects.all().order_by("name")

    if request.method == 'POST':
        equipment_index = request.POST.get('equipment')
        type_index = request.POST.get('type')
        excel_file = request.FILES['file']
        selected_sheet = request.POST.get('selected_sheet')

        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True)
            ws = wb[selected_sheet]
        except Exception as e:
            messages.error(request, f"Error loading Excel file or sheet: {str(e)}")
            return redirect('upload_price_list')

        # Assuming headers are in the first row
        headers = [cell.value for cell in ws[1]]
        headers_lower = [header.lower() if header is not None else None for header in headers]

        required_columns = ['part no', 'price']
        for column in required_columns:
            if column not in headers_lower:
                messages.error(request, f"Column '{column}' not found in the Excel file.")
                return redirect('upload_price_list')

        supplier_index = request.POST.get('supplier')
        brand_index = request.POST.get('brand')

        # Assume headers is a list containing the column headers

        # Function to get index if header is present, otherwise None
        def get_index(header_name):
            return headers_lower.index(header_name) if header_name in headers_lower else None

        # Mapping headers to their indices
        index_mapping = {
            'product_name': get_index('part no'),
            'availability': get_index('availability'),
            'price': get_index('price'),
            'series': get_index('series'),
            'product_link': get_index('productlink'),
            'description': get_index('description'),
            'processor': get_index('processor'),
            'os': get_index('os'),
            'stock': get_index('stock'),
        }

        # Now you can access the indices using the keys
        product_name_index = index_mapping['product_name']
        price_index = index_mapping['price']
        series_index = index_mapping['series']
        product_link_index = index_mapping['product_link']
        description_index = index_mapping['description']
        processor_index = index_mapping['processor']
        os_index = index_mapping['os']
        stock_index = index_mapping['stock']
        availability_index = index_mapping['availability']

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Create instances of Supplier, Brand, and Equipment if they don't exist
            supplier = Supplier.objects.get(id=supplier_index)
            brand = Brand.objects.get(id=brand_index)
            equipment = Equipment.objects.get(id=equipment_index)

            product_name = row[product_name_index]
            price_name = row[price_index]
            if product_name is not None and price_name is not None:
                # Create an instance of PriceList
                price_list_obj = LaptopPriceList(
                    product_name=product_name,
                    price=price_name,
                    description=row[description_index] if description_index is not None else '',
                    availability=row[availability_index] if availability_index is not None else '',
                    processor=row[processor_index] if processor_index is not None else '',
                    os=row[os_index] if os_index is not None else '',
                    stock=row[stock_index] if stock_index is not None else '',
                    currency=request.POST.get('currency'),
                    supplier=supplier,
                    series=row[series_index] if series_index is not None else '',
                    ProductLink=row[product_link_index] if product_link_index is not None else '',
                    equipment=equipment,
                    brand=brand
                )
                if type_index:
                    equipment_type = Type.objects.get(id=type_index)
                    price_list_obj.type=equipment_type


                try:
                    price_list_obj.save()
                except Exception as e:
                    messages.error(request, f"Error Processing Data: {str(e)}")
                    return redirect('upload_price_list')
        messages.success(request, 'Products Uploaded successfully')
        return redirect('search_laptops')  # Redirect to a success page

    return render(request, 'prices/upload_price_list.html',
                  {'suppliers': suppliers, 'brands': brands, 'equipment': equipment})


@login_required
def upload_coloursoft_price_list(request):
    suppliers = Supplier.objects.all()
    brands = Brand.objects.all()
    equipment = Equipment.objects.all()

    if request.method == 'POST':
        excel_file = request.FILES['file']
        selected_sheet = request.POST.get('selected_sheet')

        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            ws = wb[selected_sheet]
        except Exception as e:
            messages.error(request, f"Error loading Excel file or sheet: {str(e)}")
            return redirect('upload_coloursoft_price_list')

        # Assuming headers are in the first row
        headers = [cell.value for cell in ws[1]]

        # Check if 'product_name' and 'price' are present in the headers

        # Assume headers is a list containing the column headers
        headers_lower = [header.lower() if header is not None else None for header in headers]

        # Function to get index if header is present, otherwise None
        def get_index(header_name):
            return headers_lower.index(header_name) if header_name in headers_lower else None

        required_columns = ['code', 'price']
        for column in required_columns:
            if column not in headers_lower:
                messages.error(request, f"Column '{column}' not found in the Excel file.")
                return redirect('upload_coloursoft_price_list')

        # Mapping headers to their indices
        index_mapping = {
            'code': get_index('code'),
            'hp_code': get_index('code 1'),
            'yield_no': get_index('yield'),
            'brand': get_index('brand'),
            'price': get_index('price'),
            'level_1': get_index('level 1'),
            'level_2': get_index('level 2'),
            'level_3': get_index('level 3'),
            'end_user': get_index('end user'),
        }

        # Now you can access the indices using the keys
        brand_index = index_mapping['brand']
        hp_code_index = index_mapping['hp_code']
        code_index = index_mapping['code']
        price_index = index_mapping['price']
        yield_no_index = index_mapping['yield_no']
        level_1_index = index_mapping['level_1']
        level_2_index = index_mapping['level_2']
        level_3_index = index_mapping['level_3']
        end_user_index = index_mapping['end_user']

        for row in ws.iter_rows(min_row=2, values_only=True):

            code_value = row[code_index]
            if code_value is not None:
                # Create instances of Brand if it doesn't exist
                brand = Brand.objects.get(id=brand_index)  # Assuming brand_index is already defined

                # Create an instance of ColoursoftPriceList
                price_list_obj = ColoursoftPriceList(
                    code=code_value,
                    price=row[price_index] if price_index is not None else '0',
                    yield_no=row[yield_no_index] if yield_no_index is not None else '0',
                    hp_code=row[hp_code_index] if hp_code_index is not None else '',
                    level_1=row[level_1_index] if level_1_index is not None else '0',
                    level_2=row[level_2_index] if level_2_index is not None else '0',
                    level_3=row[level_3_index] if level_3_index is not None else '0',
                    end_user=row[end_user_index] if end_user_index is not None else '0',
                    currency=request.POST.get('currency'),
                    brand=row[brand_index] if brand_index is not None else '0',
                )
                try:
                    price_list_obj.save()
                except Exception as e:
                    messages.error(request, f"Error Processing Data: {str(e)}")
                    return redirect('upload_coloursoft_price_list')

        messages.success(request, 'Products Uploaded successfully')
        return redirect('search_coloursoft')  # Redirect to a success page

    return render(request, 'prices/upload_coloursoft_price_list.html',
                  {'suppliers': suppliers, 'brands': brands, 'equipment': equipment})


@login_required
def upload_fellowes_price_list(request):
    suppliers = Supplier.objects.all()
    brands = Brand.objects.all()
    equipment = Equipment.objects.all()

    if request.method == 'POST':
        equipment_index = request.POST.get('equipment')
        excel_file = request.FILES['file']
        selected_sheet = request.POST.get('selected_sheet')

        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            ws = wb[selected_sheet]
        except Exception as e:
            messages.error(request, f"Error loading Excel file or sheet: {str(e)}")
            return redirect('upload_fellowes_price_list')

        # Assuming headers are in the first row
        headers = [cell.value for cell in ws[1]]
        headers_lower = [header.lower() if header is not None else None for header in headers]

        # Check if 'product_name' and 'price' are present in the headers
        required_columns = ['code', 'price']
        for column in required_columns:
            if column not in headers_lower:
                messages.error(request, f"Column '{column}' not found in the Excel file.")
                return redirect('upload_fellowes_price_list')

        # Assume headers is a list containing the column headers
        headers_lower = [header.lower() if header is not None else None for header in headers]

        # Function to get index if header is present, otherwise None
        def get_index(header_name):
            return headers_lower.index(header_name) if header_name in headers_lower else None

        # Mapping headers to their indices
        index_mapping = {
            'code': get_index('code'),
            'price': get_index('price'),
            'price1': get_index('price 1'),
            'price2': get_index('price 2'),
            'specification': get_index('specification'),
            'description': get_index('description'),
            'discount': get_index('discount'),
            'stock': get_index('stock'),
        }

        # Now you can access the indices using the keys
        code_index = index_mapping['code']
        price_index = index_mapping['price']
        price1_index = index_mapping['price1']
        price2_index = index_mapping['price2']
        description_index = index_mapping['description']
        specification_index = index_mapping['specification']
        discount_index = index_mapping['discount']
        stock_index = index_mapping['stock']

        for row in ws.iter_rows(min_row=2, values_only=True):

            code = row[code_index]
            price_name = row[price_index]
            if code is not None and price_name is not None:
                # Create an instance of PriceList
                price_list_obj = FellowesPricelist(
                    code=code,
                    price=price_name,
                    description=row[description_index] if description_index is not None else '',
                    price1=row[price1_index] if price1_index is not None else '',
                    price2=row[price2_index] if price2_index is not None else '',
                    specification=row[specification_index] if specification_index is not None else '',
                    stock=row[stock_index] if stock_index is not None else '',
                    discount=row[discount_index] if discount_index is not None else '',
                )
                try:
                    price_list_obj.save()
                except Exception as e:
                    messages.error(request, f"Error Processing Data: {str(e)}")
                    return redirect('upload_fellowes_price_list')
        messages.success(request, 'Products Uploaded successfully')
        return redirect('search_fellowes')  # Redirect to a success page

    return render(request, 'prices/upload_fellowes_price_list.html',
                  {'suppliers': suppliers, 'brands': brands, 'equipment': equipment})


@login_required
def dupload_price_list(request):
    if request.method == 'POST':
        form = PriceListForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = request.FILES['file']
                # Specify the engine based on the Excel file format
                engine = 'openpyxl' if excel_file.name.endswith('xlsx') else 'xlrd'
                df = pd.read_excel(excel_file, engine=engine)
                # Assuming your PriceList model has fields like 'product_name', 'price', 'currency'
                for index, row in df.iterrows():
                    LaptopPriceList.objects.create(
                        product_name=row['product_name'],
                        price=row['price'],
                        currency=row['currency']
                    )
                return redirect('view_price_list')  # Redirect to a success page
            except Exception as e:
                # Handle exceptions, log them, or display an error message
                print(f"An error occurred: {e}")
                # You might want to redirect to an error page or display an error message
    else:
        form = PriceListForm()
    return render(request, 'prices/upload_price_list.html', {'form': form})


@login_required
def view_laptop_price_list(request):
    price_list_items = LaptopPriceList.objects.all()
    for item in price_list_items:
        for item in price_list_items:
            if 1 <= item.price <= 350:
                item.price_min = (item.price + item.price * Decimal('0.08')).quantize(Decimal('0.00'),
                                                                                      rounding=ROUND_DOWN)
                item.price_max = (item.price + item.price * Decimal('0.10')).quantize(Decimal('0.00'),
                                                                                      rounding=ROUND_DOWN)
            elif 351 <= item.price <= 500:
                item.price_min = (item.price + item.price * Decimal('0.10')).quantize(Decimal('0.00'),
                                                                                      rounding=ROUND_DOWN)
                item.price_max = (item.price + item.price * Decimal('0.12')).quantize(Decimal('0.00'),
                                                                                      rounding=ROUND_DOWN)
            elif 501 <= item.price <= 800:
                item.price_min = (item.price + item.price * Decimal('0.08')).quantize(Decimal('0.00'),
                                                                                      rounding=ROUND_DOWN)
                item.price_max = (item.price + item.price * Decimal('0.10')).quantize(Decimal('0.00'),
                                                                                      rounding=ROUND_DOWN)
            elif item.price > 800:
                item.price_min = (item.price + item.price * Decimal('0.06')).quantize(Decimal('0.00'),
                                                                                      rounding=ROUND_DOWN)
                item.price_max = (item.price + item.price * Decimal('0.08')).quantize(Decimal('0.00'),
                                                                                      rounding=ROUND_DOWN)

    return render(request, 'prices/view_price_list.html', {'price_list_items': price_list_items})


@login_required
def create_supplier(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            Supplier.objects.create(name=name)
            return redirect('list-suppliers')  # Redirect to a list view or any other page
        else:
            return HttpResponse("Invalid data submitted for creating a supplier.")

    return render(request, 'prices/create_supplier.html')


@login_required
def create_brand(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            Brand.objects.create(name=name)
            return redirect('list_brands')  # Redirect to a list view or any other page
        else:
            return HttpResponse("Invalid data submitted for creating a supplier.")

    return render(request, 'prices/create_brand.html')


def min_price(price, equipment_type, item_currency):
    type = str(equipment_type)
    currency = str(item_currency)
    exchange_rate = Exchange.objects.first().rate

    if currency == "KES":
        price2 = price / exchange_rate
    else:
        price2 = price

    if type in ["Laptop", "DESKTOP"]:
        if 1 <= price2 <= 350:
            return (price + price * Decimal('0.12')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif 351 <= price <= 500:
            return (price + price * Decimal('0.10')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif 501 <= price2 <= 800:
            return (price + price * Decimal('0.08')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif price2 > 800:
            return (price + price * Decimal('0.06')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["PROJECTORS"]:
        return (price + price * Decimal('0.10')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["PRINTERS & SCANNERS"]:

        if 1 <= price2 <= 50:
            return (price + price * Decimal('0.25')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif 51 <= price2 <= 100:
            return (price + price * Decimal('0.20')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif 101 <= price2 <= 250:
            return (price + price * Decimal('0.12')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif price2 > 251:
            return (price + price * Decimal('0.10')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["SERVER PARTS & ACCESSORIES"]:

        if 1 <= price2 <= 200:
            return (price + price * Decimal('0.30')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif price2 > 201:
            return (price + price * Decimal('0.25')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["APPLE iPad/ MacBook"]:

        if 1 <= price <= 600:
            return (price + price * Decimal('0.15')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif price > 601:
            return (price + price * Decimal('0.12')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["NETWORK ITEMS /DLINK/ TPLINK"]:

        if 1 <= price2 <= 100:
            if currency == "USD":
                # Convert the amount to USD
                amount_in_usd = 2000 / exchange_rate
                # Round the result to two decimal places
                amount_in_usd = round(amount_in_usd, 2)
                return (price + amount_in_usd)
            else:
                return (price + 2000)
        elif price2 > 101:
            return (price + price * Decimal('0.30')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["UBIQUITY/ MIKROTIK"]:

        if 1 <= price2 <= 238.10:
            return (price + price * Decimal('0.15')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif price2 > 238.11:
            return (price + price * Decimal('0.10')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["UPS", "SOFTWARE"]:
        return (price + price * Decimal('0.12')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["CARTRIDGES & RIBBONS", "TONNERS"]:
        return (price + price * Decimal('0.15')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["ACCESSORIES"]:

        if 1 <= price2 <= 100:
            if currency == "USD":

                amount_in_usd = 2000 / exchange_rate
                # Round the result to two decimal places
                amount_in_usd = round(amount_in_usd, 2)
                return (price + amount_in_usd)
            else:
                return (price + 2000)
        elif price2 > 101:
            return (price + price * Decimal('0.30')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["PROJECTOR SCREENS"]:

        if 1 <= price2 <= 190.48:
            return (price + price * Decimal('0.30')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif price2 > 190.49:
            return (price + price * Decimal('0.15')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["CCTV PRODUCTS"]:

        if 1 <= price2 <= 100:
            if currency == "USD":

                amount_in_usd = 2000 / exchange_rate
                # Round the result to two decimal places
                amount_in_usd = round(amount_in_usd, 2)
                return (price + amount_in_usd)
            else:
                return (price + 2000)
        elif 101 <= price2 <= 150:
            return (price + price * Decimal('0.20')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif 151 <= price2 <= 250:
            return (price + price * Decimal('0.15')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif 251 <= price2 <= 500:
            return (price + price * Decimal('0.10')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif price2 > 501:
            return (price + price * Decimal('0.8')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["WORKSTATION"]:
        return (price + price * Decimal('0.08')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["APPLE ACCESSORIES"]:
        return (price + price * Decimal('0.30')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["UPS BATTERY"]:
        return (price + price * Decimal('0.70')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["MEMORIES"]:
        return (price + price * Decimal('0.40')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["HARD DRIVES"]:
        return (price + price * Decimal('0.25')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["SOPHOS"]:
        return (price + price * Decimal('0.25')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["PARTS (BATTERY/KEYBOARD)"]:
        return (price + price * Decimal('0.90')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["POS PRODUCTS"]:
        return (price + price * Decimal('0.20')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["SERVERS"]:
        if 1 <= price2 <= 1000:
            return (price + price * Decimal('0.25')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif 1001 <= price2 <= 5000:
            return (price + price * Decimal('0.15')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif price2 > 5001:
            return (price + price * Decimal('0.13')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
    else:
        return 0.00  # No modification for other equipment types


def max_price(price, equipment_type, item_currency):
    type = str(equipment_type)
    currency = str(item_currency)
    exchange_rate = Exchange.objects.first().rate
    if currency == "KES":
        price2 = price / exchange_rate
    else:
        price2 = price

    if type in ["Laptop", "DESKTOP"]:
        if 1 <= price2 <= 350:
            return (price + price * Decimal('0.14')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif 351 <= price2 <= 500:
            return (price + price * Decimal('0.12')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif 501 <= price2 <= 800:
            return (price + price * Decimal('0.10')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif price2 > 800:
            return (price + price * Decimal('0.08')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["PROJECTORS"]:
        return (price + price * Decimal('0.15')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["PARTS (BATTERY/KEYBOARD)"]:
        return (price + price)

    elif type in ["SOPHOS"]:
        return (price + price * Decimal('0.30')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["POS PRODUCTS"]:
        return (price + price * Decimal('0.25')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["UPS BATTERY"]:
        return (price + price * Decimal('0.75')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["MEMORIES"]:
        return (price + price * Decimal('0.50')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["CARTRIDGES & RIBBONS", "TONNERS"]:
        return (price + price * Decimal('0.18')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["HARD DRIVES"]:
        return (price + price * Decimal('0.25')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["APPLE ACCESSORIES"]:
        return (price + price * Decimal('0.40')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["UPS", "SOFTWARE"]:
        return (price + price * Decimal('0.15')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["WORKSTATION"]:
        return (price + price * Decimal('0.10')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["PROJECTOR SCREENS"]:

        if 1 <= price2 <= 190.48:
            return (price + price * Decimal('0.40')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif price2 > 190.49:
            return (price + price * Decimal('0.20')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["UBIQUITY/ MIKROTIK"]:

        if 1 <= price2 <= 238.10:
            return (price + price * Decimal('0.20')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif price2 > 238.11:
            return (price + price * Decimal('0.15')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["NETWORK ITEMS /DLINK/ TPLINK"]:

        if 1 <= price2 <= 100:
            if currency == "USD":
                amount_in_usd = 2000 / exchange_rate
                # Round the result to two decimal places
                amount_in_usd = round(amount_in_usd, 2)
                return (price + amount_in_usd)
            else:
                return (price + 2000)
        elif price2 > 101:
            return (price + price * Decimal('0.40')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["APPLE iPad/ MacBook"]:

        if 1 <= price2 <= 600:
            return (price + price * Decimal('0.20')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif price2 > 601:
            return (price + price * Decimal('0.15')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["PRINTERS & SCANNERS"]:

        if 1 <= price2 <= 50:
            return (price + price * Decimal('0.30')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif 51 <= price2 <= 100:
            return (price + price * Decimal('0.25')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif 101 <= price2 <= 250:
            return (price + price * Decimal('0.15')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif price2 > 251:
            return (price + price * Decimal('0.12')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["SERVER PARTS & ACCESSORIES"]:

        if 1 <= price2 <= 200:
            return (price + price * Decimal('0.35')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif price2 > 201:
            return (price + price * Decimal('0.30')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["CCTV PRODUCTS"]:

        if 1 <= price2 <= 100:
            if currency == "USD":
                amount_in_usd = 2000 / exchange_rate
                # Round the result to two decimal places
                amount_in_usd = round(amount_in_usd, 2)
                return (price + amount_in_usd)
            else:
                return (price + 2000)
        elif 101 <= price2 <= 150:
            return (price + price * Decimal('0.25')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif 151 <= price2 <= 250:
            return (price + price * Decimal('0.20')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif 251 <= price2 <= 500:
            return (price + price * Decimal('0.15')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif price2 > 501:
            return (price + price * Decimal('0.10')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
    elif type in ["ACCESSORIES"]:
        if 1 <= price2 <= 100:
            if currency == "USD":
                amount_in_usd = 2000 / exchange_rate
                # Round the result to two decimal places
                amount_in_usd = round(amount_in_usd, 2)
                return (price + amount_in_usd)
            else:
                return (price + 2000)
        elif price2 > 101:
            return (price + price * Decimal('0.40')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)

    elif type in ["SERVERS"]:

        if 1 <= price2 <= 1000:
            return (price + price * Decimal('0.30')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif 1001 <= price2 <= 5000:
            return (price + price * Decimal('0.20')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
        elif price2 > 5001:
            return (price + price * Decimal('0.15')).quantize(Decimal('0.00'), rounding=ROUND_DOWN)
    else:
        return 0.00  # No modification for other equipment types


@login_required
def create_equipment(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        suppliers_ids = request.POST.getlist('suppliers')  # Assuming suppliers are selected through checkboxes

        if name:
            equipment = Equipment.objects.create(name=name)

            if suppliers_ids:
                equipment.suppliers.set(suppliers_ids)

            return redirect('list-equipment')  # Redirect to a list view or any other page
        else:
            return HttpResponse("Invalid data submitted for creating an equipment.")

    suppliers = Supplier.objects.all()
    return render(request, 'prices/create_equipment.html', {'suppliers': suppliers})


@login_required
def list_suppliers(request):
    suppliers = Supplier.objects.all().order_by('name')
    return render(request, 'prices/list_suppliers.html', {'suppliers': suppliers})


@login_required
def list_equipment(request):
    equipment = Equipment.objects.all().order_by('name')
    return render(request, 'prices/list_equipment.html', {'equipment': equipment})


@login_required
def list_brands(request):
    brands = Brand.objects.all().order_by('name')
    return render(request, 'prices/list_brands.html', {'brands': brands})


@login_required
def search_laptops(request):
    query = request.GET.get('q', '').lower().strip()
    selected_fields = request.GET.getlist('fields', [])

    # Define the fields you want to allow searching
    allowed_fields = ['processor', 'brand__name', 'series', 'description', 'product_name']

    # Create a dictionary to map the field names to their corresponding lookup
    field_lookup = {
        'product_name': 'icontains',
        'processor': 'icontains',
        'brand__name': 'icontains',
        'series': 'icontains',
        'description': 'icontains',  # Use icontains for case-insensitive search
        # Add other fields as needed
    }
    equipment_id = request.GET.get('equipment')
    type_id = request.GET.get('type')
    currency = request.GET.get('currency')
    # Build the Q objects for the selected fields
    q_objects = Q()
    for field in selected_fields:
        if field in allowed_fields:
            lookup = field_lookup[field]
            q_objects |= Q(**{f'{field}__{lookup}': query})

    # Check if both query and selected_fields are empty
    if not query and not selected_fields:
        # Return an empty queryset or default products if needed
        laptops = LaptopPriceList.objects.none()
    # Additional filter for the description field
    elif 'description' in selected_fields and query:
        # Split the query into keywords using both semicolons and spaces
        keywords = [kw.strip().lower() for kw in re.split(r'[;\s]+', query)]

        # Construct a list of Q objects for each keyword
        keyword_queries = [Q(description__icontains=keyword) for keyword in keywords]
        laptops = LaptopPriceList.objects.all()
        # Combine the Q objects using the | operator
        laptops = laptops.filter(*keyword_queries)
        if equipment_id:
            laptops = laptops.filter(equipment=equipment_id)
            if type_id:
                laptops = laptops.filter(type=type_id)

    else:
        # Query the model using the constructed Q objects
        laptops = LaptopPriceList.objects.filter(q_objects)
        if equipment_id:
            laptops = laptops.filter(equipment=equipment_id)
            if type_id:
                laptops = laptops.filter(type=type_id)
        # Filter by equipment if selected

        if equipment_id:
            laptops = laptops.filter(equipment=equipment_id)
            if type_id:
                laptops = laptops.filter(type=type_id)

    # Retrieve all equipment for the dropdown
    all_equipment = Equipment.objects.all()
    all_types = Type.objects.all()

    for item in laptops:
        exchange_rate = Exchange.objects.first().rate
        exchange_rate2 = Exchange.objects.first().rate2

        if item.currency == "KES":
            price2 = item.price / exchange_rate
        else:
            price2 = item.price

        item.price_min = calculate_discounted_price(item.price, item.equipment, price2)
        item.price_max = calculate_discounted_price2(item.price, item.equipment, price2)

        if currency == "KES" and item.currency == "USD":

            item.price = item.price * exchange_rate2
            item.price_max = item.price_max * exchange_rate2
            item.price_min = item.price_min * exchange_rate2
            item.currency = "KES"

        elif currency == "USD" and item.currency == "KES":

            item.currency = "USD"

            item.price = item.price / exchange_rate
            item.price_max = item.price_max / exchange_rate
            item.price_min = item.price_min / exchange_rate

    context = {'laptops': laptops, 'query': query, 'selected_fields': selected_fields, 'allowed_fields': allowed_fields,
               'all_equipment': all_equipment,'all_types':all_types }
    return render(request, 'prices/search_laptops.html', context)


@login_required
def search_coloursoft(request):
    query = request.GET.get('q', '').lower().strip()
    selected_fields = request.GET.getlist('fields', [])

    # Define the fields you want to allow searching
    allowed_fields = ['brand', 'code', 'yield_no', 'hp_code']

    # Create a dictionary to map the field names to their corresponding lookup
    field_lookup = {
        'code': 'icontains',
        'brand': 'icontains',
        'yield_no': 'icontains',
        'hp_code': 'icontains',
    }

    equipment_id = request.GET.get('equipment')
    # Build the Q objects for the selected fields
    q_objects = Q()
    for field in selected_fields:
        if field in allowed_fields:
            lookup = field_lookup[field]
            q_objects |= Q(**{f'{field}__{lookup}': query})

    # Check if both query and selected_fields are empty
    if not query and not selected_fields:
        # Return an empty queryset or default products if needed
        laptops = ColoursoftPriceList.objects.none()
    # Additional filter for the description field
    elif 'code' in selected_fields and query:
        # Split the query into keywords using both semicolons and spaces
        keywords = [kw.strip().lower() for kw in re.split(r'[;\s]+', query)]

        # Construct a list of Q objects for each keyword
        keyword_queries = [Q(code__icontains=keyword) for keyword in keywords]
        laptops = ColoursoftPriceList.objects.all()
        # Combine the Q objects using the | operator
        laptops = laptops.filter(*keyword_queries)
        if equipment_id:
            laptops = laptops.filter(equipment=equipment_id)

    else:
        # Query the model using the constructed Q objects
        laptops = ColoursoftPriceList.objects.filter(q_objects)
        if equipment_id:
            laptops = laptops.filter(equipment=equipment_id)

        # Filter by equipment if selected

        if equipment_id:
            laptops = laptops.filter(equipment=equipment_id)

    # Retrieve all equipment for the dropdown
    all_equipment = Equipment.objects.all()

    context = {'laptops': laptops, 'query': query, 'selected_fields': selected_fields, 'allowed_fields': allowed_fields,
               'all_equipment': all_equipment, }
    return render(request, 'prices/search_coloursoft.html', context)


@login_required
def search_fellowes(request):
    query = request.GET.get('q', '').lower().strip()
    selected_fields = request.GET.getlist('fields', [])

    # Define the fields you want to allow searching
    allowed_fields = ['description', 'code', 'specification']

    # Create a dictionary to map the field names to their corresponding lookup
    field_lookup = {
        'code': 'icontains',
        'description': 'icontains',
        'specification': 'icontains',
    }

    equipment_id = request.GET.get('equipment')
    # Build the Q objects for the selected fields
    q_objects = Q()
    for field in selected_fields:
        if field in allowed_fields:
            lookup = field_lookup[field]
            q_objects |= Q(**{f'{field}__{lookup}': query})

    # Check if both query and selected_fields are empty
    if not query and not selected_fields:
        # Return an empty queryset or default products if needed
        laptops = FellowesPricelist.objects.none()
    # Additional filter for the description field
    elif 'description' in selected_fields and query:
        # Split the query into keywords using both semicolons and spaces
        keywords = [kw.strip().lower() for kw in re.split(r'[;\s]+', query)]

        # Construct a list of Q objects for each keyword
        keyword_queries = [Q(description__icontains=keyword) for keyword in keywords]
        laptops = FellowesPricelist.objects.all()
        # Combine the Q objects using the | operator
        laptops = laptops.filter(*keyword_queries)


    else:
        # Query the model using the constructed Q objects
        laptops = FellowesPricelist.objects.filter(q_objects)
        if equipment_id:
            laptops = laptops.filter(equipment=equipment_id)

    # Retrieve all equipment for the dropdown
    all_equipment = Equipment.objects.all()

    context = {'laptops': laptops, 'query': query, 'selected_fields': selected_fields, 'allowed_fields': allowed_fields,
               'all_equipment': all_equipment, }
    return render(request, 'prices/search_fellowes.html', context)


@login_required
def edit_exchange(request):
    exchange, created = Exchange.objects.get_or_create(pk=1)  # Adjust the condition based on your requirements

    if request.method == 'POST':
        rate = request.POST.get('rate')
        rate2 = request.POST.get('rate2')
        try:
            rate2 = Decimal(rate2)
            rate = Decimal(rate)
            exchange.rate2 = rate2
            exchange.rate = rate
            exchange.save()
        except ValueError:
            # Handle invalid input (non-numeric)
            pass

    return render(request, 'prices/edit_exchange.html', {'exchange': exchange})


@login_required
def price_rules_for_equipment(request, equipment_id):
    equipment = get_object_or_404(Equipment, pk=equipment_id)
    price_rules = PriceRule.objects.filter(product_type=equipment)

    if request.method == 'POST':
        # Assuming you have form data in the request.POST dictionary
        # Extract data and update PriceRule objects accordingly
        for price_rule in price_rules:
            # Use the appropriate field names based on your form
            discount_percentage = Decimal(request.POST.get(f"discount_percentage_{price_rule.id}", 0))
            discount_percentage2 = Decimal(request.POST.get(f"discount_percentage2_{price_rule.id}", 0))
            constant = request.POST.get(f"constant_{price_rule.id}")
            if constant == "on":
                constant = True
            else:
                constant = False
            price_rule.discount_percentage = discount_percentage
            price_rule.discount_percentage2 = discount_percentage2

            price_rule.constant = constant

            price_rule.save()

    return render(request, 'prices/edit_price_rules.html', {'equipment': equipment, 'price_rules': price_rules})

def add_price_rule(request,equipment_id):
    if request.method == 'POST':
        # Retrieve form data from the request
        price_range_start_data = request.POST.getlist('price_range_startb[]')
        constant = request.POST.getlist('constantb[]')
        price_range_end_data = request.POST.getlist('price_range_endb[]')
        discount_percentage_data = request.POST.getlist('discount_percentageb[]')
        discount_percentage2_data = request.POST.getlist('discount_percentage2b[]')
        equipment_id =equipment_id
        print("data")
        print(price_range_start_data)
        print(constant)
        for i in range(len(price_range_start_data)):
            if (price_range_start_data[i]):
                price_rule = PriceRule(
                    product_type_id=equipment_id,
                    price_range_start=Decimal(price_range_start_data[i]),
                    discount_percentage=Decimal(discount_percentage_data[i]),
                    discount_percentage2=Decimal(discount_percentage2_data[i]),

                )

                if price_range_end_data[i]:
                    price_rule.price_range_end = price_range_end_data[i]

                price_rule.save()

        return redirect('price_rules_for_equipment',equipment_id)  # Change 'success_page' to the actual URL name or path

    equipment_types = Equipment.objects.all()  # Assuming Equipment is the related model
    return render(request, 'prices/edit_price_rules.html', {'equipment_types': equipment_types})

@login_required
def type_list(request):
    types = Type.objects.all()
    return render(request, 'prices/type_list.html', {'types': types})


@login_required
def create_type(request):
    equipments = Equipment.objects.all()
    if request.method == 'POST':
        name = request.POST.get('name')
        equipment_id = request.POST.get('equipment_id')  # Assuming you have a form field for selecting equipment
        equipment = get_object_or_404(Equipment, pk=equipment_id)
        Type.objects.create(name=name, equipment=equipment)
        messages.success(request, 'Type Created successfully')
        return redirect('type_list')  # Redirect to the list view after creating a type
    return render(request, 'prices/create_type.html', {'equipments': equipments})


@login_required
def edit_type(request, type_id):
    type_instance = get_object_or_404(Type, pk=type_id)
    equipments = Equipment.objects.all()

    if request.method == 'POST':
        name = request.POST.get('name')
        equipment_id = request.POST.get('equipment_id')  # Assuming you have a form field for selecting equipment
        equipment = get_object_or_404(Equipment, pk=equipment_id)

        type_instance.name = name
        type_instance.equipment = equipment
        type_instance.save()
        messages.success(request, 'Type Edited successfully')
        return redirect('type_list')  # Redirect to the list view after editing a type

    return render(request, 'prices/edit_type.html', {'type_instance': type_instance, 'equipments': equipments})


@login_required
def delete_type(request, type_id):
    type_instance = get_object_or_404(Type, pk=type_id)

    if request.method == 'POST':
        type_instance.delete()
        messages.warning(request, 'Type Deleted successfully')
        return redirect('type_list')  # Redirect to the list view after deleting a type

    return render(request, 'prices/delete_type.html', {'type_instance': type_instance})

def get_types(request):
    if request.method == 'GET':
        equipment_id = request.GET.get('equipment_id')

        # Assuming you have a ForeignKey relationship between Type and Equipment
        types = Type.objects.filter(equipment_id=equipment_id).values('id', 'name')

        return JsonResponse({'types': list(types)})

    return JsonResponse({'error': 'Invalid request method'}, status=400)

@login_required
def delete_laptop_price(request, laptop_price_id):
    laptop_price = get_object_or_404(LaptopPriceList, pk=laptop_price_id)

    if request.method == 'POST':
        laptop_price.delete()
        messages.warning(request, 'Product Deleted successfully')
        return redirect('search_laptops')  # Redirect to the list view after deleting a laptop price

    return render(request, 'prices/delete_laptop_price.html', {'laptop_price': laptop_price})

@login_required
def edit_laptop_price(request, laptop_price_id):
    laptop_price = get_object_or_404(LaptopPriceList, pk=laptop_price_id)
    suppliers = Supplier.objects.all()
    equipments = Equipment.objects.all()
    brands = Brand.objects.all()
    types = Type.objects.all()
    print("data")
    print(laptop_price.supplier_id )
    print(request.POST.get('supplier'))
    if request.method == 'POST':
        print(laptop_price.supplier_id)
        print(request.POST.get('supplier'))
        # Handle form data and update the laptop_price instance
        laptop_price.product_name = request.POST.get('product_name')
        laptop_price.type_id = request.POST.get('type')
        laptop_price.supplier_id = request.POST.get('supplier')
        laptop_price.series = request.POST.get('series')
        laptop_price.ProductLink = request.POST.get('ProductLink')
        laptop_price.equipment_id = request.POST.get('equipment')
        laptop_price.brand_id = request.POST.get('brand')
        laptop_price.price = request.POST.get('price')
        laptop_price.description = request.POST.get('description')
        laptop_price.os = request.POST.get('os')
        laptop_price.stock = request.POST.get('stock')
        laptop_price.currency = request.POST.get('currency')
        # Update other fields as needed
        laptop_price.save()
        messages.success(request, 'Product Edited successfully')
        return redirect('search_laptops')  # Redirect to the list view after editing a laptop price

    return render(request, 'prices/edit_laptop_price.html', {'laptop_price': laptop_price,'suppliers':suppliers,'equipments':equipments,'brands':brands,'types':types})

@csrf_exempt
def search_laptops_js(request):
    if request.method == 'POST':
        query = request.POST.get('q', '')
        equipment = request.POST.get('equipment', '')
        fields = request.POST.get('fields', '')
        currency = request.POST.get('currency', '')

        # Perform the search based on the form data
        laptops = LaptopPriceList.objects.filter(
            product_name__icontains=query,
            # Add other filters based on the form data
        )

        # Prepare JSON data
        laptop_data = []
        for laptop in laptops:
            laptop_data.append({
                'product_name': laptop.product_name,
                'supplier': laptop.supplier.name,
                'price': str(laptop.price),
                'description': laptop.description,
                # Add other fields as needed
            })

        # Return JSON response
        return JsonResponse(laptop_data, safe=False)
    else:
        return JsonResponse({'error': 'Invalid request method'})